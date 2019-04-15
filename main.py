"""
Copyright 2019 DevHyung
   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at
       http://www.apache.org/licenses/LICENSE-2.0
   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""
"""
This file is manual version.
Auto version To Be Announced 
"""
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver
from bs4 import BeautifulSoup
import os
import time

class ExcelDriver:
    def __init__(self, _fileName,_header):
        self.fileName = _fileName + ".xlsx"
        self.header = _header
        self.create_File()
    def create_File(self):
        if  os.path.exists(self.fileName):
            self.log('i','exist file name ')
        else:  # 새로만드는건
            book = Workbook()
            sheet = book.active
            sheet.title = 'default'
            sheet.append(self.header)
            # if modify cell width, write down
            #['저자', '내용', '좋아요/조회수', '해쉬태그', '댓글']
            sheet.column_dimensions['A'].width = 20
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 20
            book.save(self.fileName)

    def append_data(self,_data):
        '''

        :param _data: 1-d list type input data to excel ,
        :return:
        '''
        book = load_workbook(self.fileName)
        sheet = book.active
        sheet.append(_data)
        book.save(self.fileName)
        self.log('s',"데이터 저장완료.")

    def append_data_list(self,_dataList):
        '''
        :param _dataList: 2-d list type input data to excel ,
        :return:
        '''
        book = load_workbook(self.fileName)
        sheet = book.active
        for data in _dataList:
            sheet.append(data)
        book.save(self.fileName)
        self.log('s', "{}개 데이터 저장완료.".format(len(_dataList)))

    @staticmethod
    def log(tag, text):
        # Info tag
        if (tag == 'i'):
            print("[INFO] " + text)
        # Error tag
        elif (tag == 'e'):
            print("[ERROR] " + text)
        # Success tag
        elif (tag == 's'):
            print("[SUCCESS] " + text)
if __name__ == "__main__":
    ''' CONFIG AREA '''
    FILENAME = "인스타그램" # without file extention
    headerList = ['저자', '내용', '좋아요/조회수', '해쉬태그', '댓글'] #엑셀 맨위 머릿글
    COMMENT_MAX = 5 # 5개까지만 가져온다
    
    ''' Run '''
    excel = ExcelDriver(FILENAME,headerList) # excel init

    driver = webdriver.Chrome('./chromedriver')
    driver.get('https://www.instagram.com/hurom.korea/')
    while True:
        enter = input(">>> 파싱하시려는 게시물에서 엔터를 눌러주세요 (종료는 x):").strip()
        if enter.lower() == 'x':
            break

        bs = BeautifulSoup(driver.page_source, 'lxml')

        postE = bs.find('li', {'role': "menuitem"})
        aTags = postE.find('span').find_all('a')
        likeSection = bs.find_all('section')[3]
        # result variable
        tagList = []
        commentList = []
        author = postE.find('h2').get_text().strip()
        content = postE.find('span').get_text()
        viewOrLike = likeSection.get_text().strip()

        for a in aTags:
            tag = a.get_text().strip()
            tagList.append(tag)
            content = content.replace(tag, '', 1)
        content = content.strip()

        commentLis = bs.find_all('li', {'role': 'menuitem'})[1:COMMENT_MAX+1]
        for comment in commentLis:
            commentList.append(comment.find('span').get_text())
        excel.append_data([author,content,viewOrLike,'\n'.join(tagList)]+commentList)
    driver.quit()
